import os
import openpyxl
from openpyxl.drawing.image import Image
import io
from PIL import Image as PILImage
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import string

# 设置PIL图像大小限制，避免解压炸弹攻击的限制
# 注意：仅在确认Excel文件安全的情况下使用
PILImage.MAX_IMAGE_PIXELS = None

def get_column_letter(col_idx):
    """将列索引转换为Excel列字母"""
    return string.ascii_uppercase[col_idx]

def get_column_index(col_letter):
    """将Excel列字母转换为索引"""
    return string.ascii_uppercase.index(col_letter)

def extract_images_from_excel(excel_file, output_folder='extracted_images', name_col_idx=1, img_col_indices=None, name_suffix=''):
    """从Excel文件中提取图片
    
    参数:
        excel_file: Excel文件路径
        output_folder: 输出文件夹路径
        name_col_idx: 用于命名图片的列索引（从0开始）
        img_col_indices: 包含图片的列索引列表（从0开始）
    """
    # 如果未指定图片列，默认使用I和J列
    if img_col_indices is None:
        img_col_indices = [8, 9]  # I列=8, J列=9 (A=0, B=1, ...)
    
    # 创建输出文件夹（如果不存在）
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"创建输出文件夹: {output_folder}")
    
    # 加载Excel工作簿
    print(f"正在打开Excel文件: {excel_file}")
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        error_msg = f"错误: 无法打开Excel文件: {e}"
        print(error_msg)
        return False, error_msg
    
    # 获取活动工作表
    ws = wb.active
    print(f"正在处理工作表: {ws.title}")
    
    # 查找所有图片
    images_extracted = 0
    
    # 检查是否有图片对象
    if not hasattr(ws, '_images') or len(ws._images) == 0:
        error_msg = "警告: 工作表中没有找到图片对象!\n提示: 请确认Excel文件中确实包含图片，而不是链接或其他对象。"
        print(error_msg)
        return False, error_msg
    
    print(f"在工作表中找到 {len(ws._images)} 个图片对象")
    
    # 遍历所有图片对象
    for idx, image in enumerate(ws._images):
        try:
            # 获取图片的单元格位置
            col = image.anchor._from.col
            row = image.anchor._from.row + 1  # 行号从0开始，加1使其与Excel显示一致
            
            print(f"处理第 {idx+1} 个图片对象 (位置: 行 {row}, 列 {get_column_letter(col)}列)")
            
            # 检查是否在指定的图片列中
            if col in img_col_indices:
                column_letter = get_column_letter(col)
                
                # 获取命名列对应行的值作为文件名
                cell_value = ws.cell(row=row, column=name_col_idx+1).value  # 列索引从0开始，但cell()方法从1开始
                
                if cell_value:
                    # 清理文件名（移除不允许的字符）
                    filename = str(cell_value).strip()
                    # 替换文件名中的非法字符
                    for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                        filename = filename.replace(char, '_')
                    
                    # 添加后缀和列标识到文件名
                    if name_suffix:
                        filename = f"{filename}{name_suffix}_{column_letter}"
                    else:
                        filename = f"{filename}_{column_letter}"
                    
                    try:
                        # 获取图片数据
                        img_data = image._data()
                        
                        # 使用PIL打开图片以确定格式
                        img = PILImage.open(io.BytesIO(img_data))
                        img_format = img.format.lower() if img.format else 'png'
                        
                        # 保存图片
                        img_path = os.path.join(output_folder, f"{filename}.{img_format}")
                        with open(img_path, 'wb') as f:
                            f.write(img_data)
                        
                        print(f"已保存图片: {img_path} (来自行 {row}, {column_letter}列)")
                        images_extracted += 1
                    except Exception as e:
                        print(f"错误: 处理图片时出错: {e}")
                else:
                    print(f"警告: 行 {row} 的命名列没有值，跳过该图片")
            else:
                print(f"跳过非指定列的图片 (行 {row}, 列 {get_column_letter(col)})")
        except Exception as e:
            print(f"错误: 处理图片对象时出错: {e}")
    
    result_msg = f"完成! 共提取了 {images_extracted} 张图片到 {output_folder} 文件夹"
    print(result_msg)
    if images_extracted == 0:
        warning_msg = "警告: 没有提取到任何图片! 请检查Excel文件格式是否正确。\n提示: 确保图片位于选定的列，且命名列包含有效的文件名。"
        print(warning_msg)
        return False, warning_msg
    
    return True, result_msg

class ExcelImageExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel图片提取工具")
        self.root.geometry("600x450")
        self.root.resizable(True, True)
        
        # 设置样式
        self.style = ttk.Style()
        self.style.configure('TButton', font=('SimHei', 10))
        self.style.configure('TLabel', font=('SimHei', 10))
        self.style.configure('TFrame', background='#f0f0f0')
        
        # 创建主框架
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Excel文件选择
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(file_frame, text="Excel文件:").pack(side=tk.LEFT, padx=5)
        self.file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path_var, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(file_frame, text="浏览...", command=self.browse_file).pack(side=tk.LEFT, padx=5)
        
        # 输出文件夹选择
        output_frame = ttk.Frame(main_frame)
        output_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(output_frame, text="输出文件夹:").pack(side=tk.LEFT, padx=5)
        self.output_path_var = tk.StringVar(value="extracted_images")
        ttk.Entry(output_frame, textvariable=self.output_path_var, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(output_frame, text="浏览...", command=self.browse_output_folder).pack(side=tk.LEFT, padx=5)
        
        # 列选择框架
        columns_frame = ttk.LabelFrame(main_frame, text="列选择", padding="5")
        columns_frame.pack(fill=tk.X, pady=10)
        
        # 命名列选择
        name_col_frame = ttk.Frame(columns_frame)
        name_col_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(name_col_frame, text="命名列:").pack(side=tk.LEFT, padx=5)
        self.name_col_var = tk.StringVar(value="B")
        name_col_combo = ttk.Combobox(name_col_frame, textvariable=self.name_col_var, values=list(string.ascii_uppercase[:26]), width=5)
        name_col_combo.pack(side=tk.LEFT, padx=5)
        ttk.Label(name_col_frame, text="(用于命名提取的图片文件)").pack(side=tk.LEFT, padx=5)
        
        # 名称后缀输入
        suffix_frame = ttk.Frame(columns_frame)
        suffix_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(suffix_frame, text="名称后缀:").pack(side=tk.LEFT, padx=5)
        self.name_suffix_var = tk.StringVar(value="")
        ttk.Entry(suffix_frame, textvariable=self.name_suffix_var, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Label(suffix_frame, text="(可选，例如: -第一批)").pack(side=tk.LEFT, padx=5)
        
        # 图片列选择
        img_col_frame = ttk.Frame(columns_frame)
        img_col_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(img_col_frame, text="图片列:").pack(side=tk.LEFT, padx=5)
        
        # 创建多选框架
        self.img_cols_vars = {}
        checkboxes_frame = ttk.Frame(img_col_frame)
        checkboxes_frame.pack(side=tk.LEFT, padx=5, fill=tk.X)
        
        # 创建每行5个复选框的网格
        row, col = 0, 0
        for letter in string.ascii_uppercase[:26]:
            self.img_cols_vars[letter] = tk.BooleanVar(value=(letter in ["I", "J"]))
            cb = ttk.Checkbutton(checkboxes_frame, text=letter, variable=self.img_cols_vars[letter])
            cb.grid(row=row, column=col, sticky="w", padx=5)
            col += 1
            if col > 4:  # 每行5个复选框
                col = 0
                row += 1
        
        # 操作按钮
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(buttons_frame, text="提取图片", command=self.extract_images).pack(side=tk.RIGHT, padx=5)
        
        # 状态和日志区域
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 创建带滚动条的文本框
        scrollbar = ttk.Scrollbar(log_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD, yscrollcommand=scrollbar.set)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)
        
        # 初始日志信息
        self.log("欢迎使用Excel图片提取工具!\n请选择Excel文件，设置命名列和图片列，然后点击'提取图片'按钮。")
    
    def browse_file(self):
        """浏览并选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xlsm *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.log(f"已选择Excel文件: {file_path}")
    
    def browse_output_folder(self):
        """浏览并选择输出文件夹"""
        folder_path = filedialog.askdirectory(title="选择输出文件夹")
        if folder_path:
            self.output_path_var.set(folder_path)
            self.log(f"已选择输出文件夹: {folder_path}")
    
    def log(self, message):
        """向日志区域添加消息"""
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)
    
    def extract_images(self):
        """提取图片按钮点击事件"""
        # 清空日志
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state=tk.DISABLED)
        
        # 获取输入参数
        excel_file = self.file_path_var.get()
        output_folder = self.output_path_var.get()
        name_col = self.name_col_var.get()
        name_suffix = self.name_suffix_var.get()
        
        # 验证输入
        if not excel_file:
            messagebox.showerror("错误", "请选择Excel文件!")
            return
        
        if not os.path.exists(excel_file):
            messagebox.showerror("错误", f"Excel文件不存在: {excel_file}")
            return
        
        # 获取选中的图片列
        img_cols = []
        for letter, var in self.img_cols_vars.items():
            if var.get():
                img_cols.append(letter)
        
        if not img_cols:
            messagebox.showerror("错误", "请至少选择一个图片列!")
            return
        
        # 转换列字母为索引
        name_col_idx = get_column_index(name_col)
        img_col_indices = [get_column_index(col) for col in img_cols]
        
        self.log(f"开始处理Excel文件: {excel_file}")
        self.log(f"输出文件夹: {output_folder}")
        self.log(f"命名列: {name_col} (索引: {name_col_idx})")
        self.log(f"图片列: {', '.join(img_cols)} (索引: {', '.join(map(str, img_col_indices))})")
        if name_suffix:
            self.log(f"名称后缀: {name_suffix}")
        
        # 执行提取
        try:
            # 重定向标准输出到日志
            import sys
            original_stdout = sys.stdout
            
            class LogRedirector:
                def __init__(self, log_func):
                    self.log_func = log_func
                
                def write(self, message):
                    if message.strip():
                        self.log_func(message.strip())
                
                def flush(self):
                    pass
            
            sys.stdout = LogRedirector(self.log)
            
            # 执行提取
            success, message = extract_images_from_excel(
                excel_file, 
                output_folder, 
                name_col_idx, 
                img_col_indices,
                name_suffix
            )
            
            # 恢复标准输出
            sys.stdout = original_stdout
            
            if success:
                messagebox.showinfo("完成", message)
            else:
                messagebox.showwarning("警告", message)
                
        except Exception as e:
            # 恢复标准输出
            sys.stdout = original_stdout
            error_message = f"处理过程中发生错误: {str(e)}"
            self.log(error_message)
            messagebox.showerror("错误", error_message)

if __name__ == "__main__":
    # 创建GUI应用
    root = tk.Tk()
    app = ExcelImageExtractorApp(root)
    root.mainloop()